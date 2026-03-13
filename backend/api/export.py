"""
Export API - Download extraction results in various formats
"""

import csv
import io
import json
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse, JSONResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from core.database import get_db, ExtractionResult, Document, Schema
from utils.logger import get_logger

logger = get_logger(__name__)
router = APIRouter()


@router.get("/{document_id}/json")
async def export_json(document_id: str, db: AsyncSession = Depends(get_db)):
    """Export extraction results as JSON."""
    results = await _get_results(document_id, db)
    doc = await _get_doc(document_id, db)

    output = {
        "document": doc.original_filename if doc else document_id,
        "exported_at": __import__("datetime").datetime.utcnow().isoformat(),
        "results": [r.extracted_data for r in results],
    }

    content = json.dumps(output, indent=2, default=str)
    return StreamingResponse(
        io.StringIO(content),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename=extraction_{document_id}.json"},
    )


@router.get("/{document_id}/csv")
async def export_csv(document_id: str, db: AsyncSession = Depends(get_db)):
    """Export extraction results as CSV (flattened)."""
    results = await _get_results(document_id, db)
    if not results:
        raise HTTPException(404, "No extraction results found")

    output = io.StringIO()
    all_rows = []

    for r in results:
        flat = _flatten_dict(r.extracted_data or {})
        all_rows.append(flat)

    if not all_rows:
        raise HTTPException(404, "No data to export")

    # Get all unique keys
    all_keys = list(dict.fromkeys(k for row in all_rows for k in row.keys()))
    writer = csv.DictWriter(output, fieldnames=all_keys, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(all_rows)

    output.seek(0)
    return StreamingResponse(
        output,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=extraction_{document_id}.csv"},
    )


@router.get("/{document_id}/excel")
async def export_excel(document_id: str, db: AsyncSession = Depends(get_db)):
    """Export extraction results as Excel."""
    results = await _get_results(document_id, db)
    if not results:
        raise HTTPException(404, "No extraction results found")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extraction Results"

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")

        all_rows = [_flatten_dict(r.extracted_data or {}) for r in results]
        all_keys = list(dict.fromkeys(k for row in all_rows for k in row.keys()))

        # Write headers
        for col, key in enumerate(all_keys, 1):
            cell = ws.cell(row=1, column=col, value=key)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Write data
        for row_num, row_data in enumerate(all_rows, 2):
            for col, key in enumerate(all_keys, 1):
                ws.cell(row=row_num, column=col, value=str(row_data.get(key, "")))

        # Auto-fit columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=extraction_{document_id}.xlsx"},
        )
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Install it with: pip install openpyxl")


def _flatten_dict(d: dict, prefix: str = "") -> dict:
    """Flatten nested dict for CSV/Excel export."""
    result = {}
    for key, value in d.items():
        full_key = f"{prefix}.{key}" if prefix else key
        if isinstance(value, dict):
            result.update(_flatten_dict(value, full_key))
        elif isinstance(value, list):
            result[full_key] = json.dumps(value)
        else:
            result[full_key] = value
    return result


async def _get_results(doc_id: str, db: AsyncSession) -> list:
    result = await db.execute(
        select(ExtractionResult).where(ExtractionResult.document_id == doc_id)
    )
    return result.scalars().all()


async def _get_doc(doc_id: str, db: AsyncSession):
    result = await db.execute(select(Document).where(Document.id == doc_id))
    return result.scalar_one_or_none()
